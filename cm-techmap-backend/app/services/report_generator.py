"""
CM TECHMAP — Report Generator Service
Generates PDF (via WeasyPrint) and Excel (via openpyxl) reports.
"""

import io
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger("cm_techmap.reports")

# Template directory
TEMPLATES_DIR = Path(__file__).parent.parent / "templates" / "reports"


class ReportGeneratorService:
    """
    Generates reports in PDF and Excel formats.
    Works with project/flight data from the database.
    """

    def generate_pdf(
        self,
        project_data: dict[str, Any],
        flights_data: list[dict[str, Any]],
        assets_data: list[dict[str, Any]],
        report_config: dict[str, Any] | None = None,
        buildings_data: list[dict[str, Any]] | None = None,
        analytics_data: dict[str, Any] | None = None,
        terrain_data: list[dict[str, Any]] | None = None,
    ) -> bytes:
        """
        Generate a PDF report using WeasyPrint with Jinja2 templates.
        Returns raw PDF bytes.
        """
        try:
            from weasyprint import HTML
            from jinja2 import Environment, FileSystemLoader
        except ImportError:
            logger.warning("WeasyPrint/Jinja2 not available — generating placeholder PDF")
            return self._generate_placeholder_pdf(project_data)

        config = report_config or {}
        report_profile = str(config.get("report_profile", "project_summary"))

        # Setup Jinja2
        env = Environment(
            loader=FileSystemLoader(str(TEMPLATES_DIR)),
            autoescape=True,
        )

        # Register custom filters
        env.filters["format_date"] = lambda d: (
            datetime.fromisoformat(str(d)).strftime("%d/%m/%Y") if d else "—"
        )
        env.filters["format_number"] = lambda n: f"{n:,.2f}" if n else "0"
        env.filters["format_area"] = lambda sqm: (
            f"{sqm:,.0f} m²" if sqm and sqm < 1_000_000
            else f"{sqm / 1_000_000:,.2f} km²" if sqm else "—"
        )
        env.filters["format_bytes"] = lambda b: (
            f"{b / 1024 / 1024:,.1f} MB" if b else "—"
        )
        env.filters["format_currency"] = lambda v: f"R$ {v:,.2f}" if v else "R$ 0,00"

        template = env.get_template("project_report.html")

        # Prepare context
        total_area = sum(f.get("area_coverage_sqm", 0) or 0 for f in flights_data)
        total_images = sum(f.get("images_count", 0) or 0 for f in flights_data)
        ortho_assets = [a for a in assets_data if a.get("asset_type") == "orthomosaic"]
        dsm_assets = [a for a in assets_data if a.get("asset_type") == "dsm"]

        # Compute urban analytics from buildings data
        buildings = buildings_data or []
        urban_analytics = self._compute_urban_analytics(buildings, total_area)
        mandatory_sections = self._build_mandatory_sections(
            report_profile=report_profile,
            project_data=project_data,
            flights_data=flights_data,
            urban_analytics=urban_analytics,
            analytics_data=analytics_data or {},
            buildings=buildings,
            terrain_data=terrain_data or [],
            config=config,
        )

        context = {
            "project": project_data,
            "flights": flights_data,
            "assets": assets_data,
            "ortho_assets": ortho_assets,
            "dsm_assets": dsm_assets,
            "total_area_sqm": total_area,
            "total_images": total_images,
            "total_flights": len(flights_data),
            "total_assets": len(assets_data),
            "report_date": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "config": config,
            # Urban analytics
            "buildings": buildings,
            "urban_analytics": urban_analytics,
            "analytics": analytics_data or {},
            "terrain_data": terrain_data or [],
            "mandatory_sections": mandatory_sections,
            "report_profile": report_profile,
            # Customization
            "custom_title": config.get("title"),
            "custom_logo": config.get("logo_url"),
            "include_buildings": config.get("include_buildings", True),
            "include_analytics": config.get("include_analytics", True),
            "ai_narrative": config.get("ai_narrative") or {},
        }

        html_content = template.render(**context)

        # Generate PDF
        pdf_bytes = HTML(string=html_content).write_pdf()
        logger.info(f"PDF generated: {len(pdf_bytes)} bytes")
        return pdf_bytes

    def _safe_ratio(self, numerator: float, denominator: float) -> float:
        if denominator <= 0:
            return 0.0
        return float(numerator / denominator)

    def _build_mandatory_sections(
        self,
        *,
        report_profile: str,
        project_data: dict[str, Any],
        flights_data: list[dict[str, Any]],
        urban_analytics: dict[str, Any],
        analytics_data: dict[str, Any],
        buildings: list[dict[str, Any]],
        terrain_data: list[dict[str, Any]],
        config: dict[str, Any],
    ) -> dict[str, dict[str, Any]]:
        built_area = float(
            (analytics_data.get("built_area_total_sqm") if analytics_data else 0)
            or urban_analytics.get("total_built_area_sqm", 0)
            or 0
        )
        terrain_area = float(
            (analytics_data.get("terrain_area_total_sqm") if analytics_data else 0)
            or sum(float(t.get("area_sqm", 0) or 0) for t in terrain_data)
            or 0
        )
        base_area = terrain_area if terrain_area > 0 else float(project_data.get("area_sqm", 0) or 0)

        avg_confidence = 0.0
        if buildings:
            avg_confidence = sum(float(b.get("confidence", 0) or 0) for b in buildings) / len(buildings)

        compactness_values = [float(t.get("compactness", 0) or 0) for t in terrain_data if t.get("compactness") is not None]
        avg_terrain_compactness = (
            sum(compactness_values) / len(compactness_values)
            if compactness_values else 0.0
        )

        build_occupation_ratio = self._safe_ratio(built_area, max(base_area, terrain_area, 1.0))
        green_ratio = max(0.0, 1.0 - build_occupation_ratio)

        iptu_rate_per_sqm = float(config.get("iptu_rate_per_sqm", 12.0) or 12.0)
        assumed_irregular_share = float(config.get("assumed_irregular_share", 0.25) or 0.25)
        taxable_gain_area = built_area * assumed_irregular_share
        annual_tax_gain_estimate = taxable_gain_area * iptu_rate_per_sqm

        qa_score = float(analytics_data.get("qa_score", 0) or 0)
        if qa_score <= 0:
            qa_score = (
                0.45 * min(max(avg_confidence, 0.0), 1.0)
                + 0.30 * min(max(green_ratio, 0.0), 1.0)
                + 0.25 * min(max(avg_terrain_compactness, 0.0), 1.0)
            )
        qa_threshold = float(config.get("qa_threshold", 0.80) or 0.80)

        per_property_rows = []
        for idx, b in enumerate(sorted(buildings, key=lambda x: float(x.get("area_sqm", 0) or 0), reverse=True)[:400], 1):
            area = float(b.get("area_sqm", 0) or 0)
            row = {
                "index": idx,
                "area_sqm": area,
                "perimeter_m": float(b.get("perimeter_m", 0) or 0),
                "height_m": float(b.get("height_m", 0) or 0),
                "floors_estimate": int(b.get("floors_estimate", 0) or 0),
                "building_type": str(b.get("building_type", "unknown") or "unknown"),
                "confidence": float(b.get("confidence", 0) or 0),
                "quality_score": float(b.get("quality_score", b.get("confidence", 0)) or 0),
                "estimated_iptu_annual": area * iptu_rate_per_sqm,
            }
            per_property_rows.append(row)

        sections = {
            "property_appraisal": {
                "title": "Laudo Técnico por Imóvel",
                "enabled": report_profile in {"property_appraisal", "project_summary", "custom"},
                "rows": per_property_rows,
                "totals": {
                    "total_properties": len(per_property_rows),
                    "total_built_area_sqm": built_area,
                },
            },
            "project_consolidated": {
                "title": "Consolidado Executivo do Projeto",
                "enabled": report_profile in {"project_consolidated", "project_summary", "comparison", "custom"},
                "metrics": {
                    "total_flights": len(flights_data),
                    "total_buildings": int(analytics_data.get("total_buildings", 0) or urban_analytics.get("total_buildings", 0) or 0),
                    "total_terrain_patches": int(analytics_data.get("total_terrain_patches", 0) or len(terrain_data)),
                    "built_area_total_sqm": built_area,
                    "terrain_area_total_sqm": terrain_area,
                    "occupation_ratio_pct": round(build_occupation_ratio * 100.0, 2),
                    "permeable_ratio_pct": round(green_ratio * 100.0, 2),
                },
            },
            "fiscal_revenue": {
                "title": "Impacto Fiscal e Arrecadação Potencial",
                "enabled": report_profile in {"fiscal_revenue", "project_summary", "custom"},
                "metrics": {
                    "iptu_rate_per_sqm": iptu_rate_per_sqm,
                    "assumed_irregular_share_pct": round(assumed_irregular_share * 100.0, 2),
                    "taxable_gain_area_sqm": taxable_gain_area,
                    "annual_tax_gain_estimate": annual_tax_gain_estimate,
                },
            },
            "technical_qa": {
                "title": "QA Técnico da Extração IA",
                "enabled": report_profile in {"technical_qa", "project_summary", "comparison", "custom"},
                "metrics": {
                    "qa_score": qa_score,
                    "qa_threshold": qa_threshold,
                    "qa_compliant": qa_score >= qa_threshold,
                    "avg_building_confidence": avg_confidence,
                    "avg_terrain_compactness": avg_terrain_compactness,
                    "measurement_run_id": analytics_data.get("measurement_run_id"),
                    "run_completed_at": analytics_data.get("completed_at"),
                    "project_code": project_data.get("code"),
                },
            },
        }
        return sections

    def _compute_urban_analytics(
        self,
        buildings: list[dict[str, Any]],
        total_area_sqm: float,
    ) -> dict[str, Any]:
        """Compute urban analytics from building extraction data."""
        if not buildings:
            return {}

        total_buildings = len(buildings)
        total_built_area = sum(b.get("area_sqm", 0) or 0 for b in buildings)
        heights = [b.get("height_m", 0) or 0 for b in buildings if b.get("height_m")]

        analytics = {
            "total_buildings": total_buildings,
            "total_built_area_sqm": round(total_built_area, 2),
            "built_area_pct": round(total_built_area / total_area_sqm * 100, 1) if total_area_sqm > 0 else 0,
            "green_area_pct": round(max(0, 100 - (total_built_area / total_area_sqm * 100)), 1) if total_area_sqm > 0 else 0,
            "avg_height_m": round(sum(heights) / len(heights), 1) if heights else 0,
            "max_height_m": round(max(heights), 1) if heights else 0,
            "min_height_m": round(min(heights), 1) if heights else 0,
        }

        # Height distribution
        height_ranges = {"0-5m": 0, "5-10m": 0, "10-20m": 0, "20-50m": 0, "50m+": 0}
        for h in heights:
            if h <= 5:
                height_ranges["0-5m"] += 1
            elif h <= 10:
                height_ranges["5-10m"] += 1
            elif h <= 20:
                height_ranges["10-20m"] += 1
            elif h <= 50:
                height_ranges["20-50m"] += 1
            else:
                height_ranges["50m+"] += 1
        analytics["height_distribution"] = height_ranges

        # Area distribution
        area_ranges = {"<100m²": 0, "100-500m²": 0, "500-1000m²": 0, ">1000m²": 0}
        for b in buildings:
            area = b.get("area_sqm", 0) or 0
            if area < 100:
                area_ranges["<100m²"] += 1
            elif area < 500:
                area_ranges["100-500m²"] += 1
            elif area < 1000:
                area_ranges["500-1000m²"] += 1
            else:
                area_ranges[">1000m²"] += 1
        analytics["area_distribution"] = area_ranges

        return analytics

    def generate_csv(
        self,
        project_data: dict[str, Any],
        flights_data: list[dict[str, Any]],
        assets_data: list[dict[str, Any]],
        buildings_data: list[dict[str, Any]] | None = None,
    ) -> bytes:
        """Generate a CSV report with project, flights, and buildings data."""
        import csv

        output = io.StringIO()
        writer = csv.writer(output)

        # Project header
        writer.writerow(["CM TECHMAP — Relatório CSV"])
        writer.writerow(["Projeto", project_data.get("name", "")])
        writer.writerow(["Código", project_data.get("code", "")])
        writer.writerow(["Cidade", project_data.get("city", "")])
        writer.writerow(["Data", datetime.now().strftime("%d/%m/%Y %H:%M")])
        writer.writerow([])

        # Flights
        writer.writerow(["=== VOOS ==="])
        writer.writerow(["Data", "Altitude (m)", "Overlap (%)", "Imagens", "Câmera", "GSD (cm)", "Área (m²)", "Status"])
        for f in flights_data:
            writer.writerow([
                f.get("flight_date", ""),
                f.get("altitude_m", ""),
                f.get("overlap_pct", ""),
                f.get("images_count", 0),
                f.get("camera_model", ""),
                f.get("gsd_cm", ""),
                f.get("area_coverage_sqm", ""),
                f.get("status", ""),
            ])
        writer.writerow([])

        # Assets
        writer.writerow(["=== ASSETS ==="])
        writer.writerow(["Tipo", "Arquivo", "Tamanho (MB)", "Resolução (cm)", "CRS", "COG"])
        for a in assets_data:
            size_mb = (a.get("file_size_bytes", 0) or 0) / 1024 / 1024
            writer.writerow([
                a.get("asset_type", ""),
                a.get("file_key", ""),
                f"{size_mb:.1f}",
                a.get("resolution_cm", ""),
                a.get("crs_epsg", ""),
                "Sim" if a.get("cog_validated") else "Não",
            ])

        # Buildings (if available)
        if buildings_data:
            writer.writerow([])
            writer.writerow(["=== EDIFICAÇÕES DETECTADAS ==="])
            writer.writerow(["ID", "Área (m²)", "Altura (m)", "Perímetro (m)", "Tipo"])
            for i, b in enumerate(buildings_data, 1):
                writer.writerow([
                    i,
                    f"{b.get('area_sqm', 0):.1f}",
                    f"{b.get('height_m', 0):.1f}",
                    f"{b.get('perimeter_m', 0):.1f}",
                    b.get("building_type", "residential"),
                ])

        csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
        logger.info(f"CSV generated: {len(csv_bytes)} bytes")
        return csv_bytes


    def generate_excel(
        self,
        project_data: dict[str, Any],
        flights_data: list[dict[str, Any]],
        assets_data: list[dict[str, Any]],
        report_config: dict[str, Any] | None = None,
    ) -> bytes:
        """
        Generate an Excel report using openpyxl.
        Returns raw XLSX bytes.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not available — generating placeholder")
            return b""

        wb = Workbook()

        # ── Sheet 1: Resumo do Projeto ────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Resumo"

        # Styles
        header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C9190B", end_color="C9190B", fill_type="solid")
        subheader_font = Font(name="Calibri", size=11, bold=True)
        data_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Title row
        ws_summary.merge_cells("A1:F1")
        title_cell = ws_summary["A1"]
        title_cell.value = f"CM TECHMAP — Relatório: {project_data.get('name', 'Projeto')}"
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.row_dimensions[1].height = 35

        # Project info
        info_rows = [
            ("Projeto:", project_data.get("name", "—")),
            ("Código:", project_data.get("code", "—")),
            ("Cidade:", project_data.get("city", "—")),
            ("Estado:", project_data.get("state", "—")),
            ("Status:", project_data.get("status", "—")),
            ("Responsável:", project_data.get("responsible", "—")),
            ("Data do Relatório:", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("", ""),
            ("MÉTRICAS", ""),
            ("Total de Voos:", len(flights_data)),
            ("Total de Imagens:", sum(f.get("images_count", 0) or 0 for f in flights_data)),
            ("Área Total (m²):", sum(f.get("area_coverage_sqm", 0) or 0 for f in flights_data)),
            ("Assets Gerados:", len(assets_data)),
        ]

        for i, (label, value) in enumerate(info_rows, start=3):
            ws_summary.cell(row=i, column=1, value=label).font = subheader_font
            ws_summary.cell(row=i, column=2, value=str(value)).font = data_font

        ws_summary.column_dimensions["A"].width = 22
        ws_summary.column_dimensions["B"].width = 40

        # ── Sheet 2: Voos ─────────────────────────────────────────────────────
        ws_flights = wb.create_sheet("Voos")
        flight_headers = ["Data", "Altitude (m)", "Overlap (%)", "Imagens", "Câmera", "GSD (cm)", "Área (m²)", "Status"]

        for col, header in enumerate(flight_headers, 1):
            cell = ws_flights.cell(row=1, column=col, value=header)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="212427", end_color="212427", fill_type="solid")
            cell.border = thin_border

        for row_idx, flight in enumerate(flights_data, 2):
            values = [
                flight.get("flight_date", "—"),
                flight.get("altitude_m", "—"),
                flight.get("overlap_pct", "—"),
                flight.get("images_count", 0),
                flight.get("camera_model", "—"),
                flight.get("gsd_cm", "—"),
                flight.get("area_coverage_sqm", "—"),
                flight.get("status", "—"),
            ]
            for col, val in enumerate(values, 1):
                cell = ws_flights.cell(row=row_idx, column=col, value=str(val))
                cell.font = data_font
                cell.border = thin_border

        # Auto-width
        for col in range(1, len(flight_headers) + 1):
            ws_flights.column_dimensions[get_column_letter(col)].width = 16

        # ── Sheet 3: Assets ───────────────────────────────────────────────────
        ws_assets = wb.create_sheet("Assets")
        asset_headers = ["Tipo", "Arquivo", "Tamanho (MB)", "Resolução (cm)", "CRS", "COG Válido"]

        for col, header in enumerate(asset_headers, 1):
            cell = ws_assets.cell(row=1, column=col, value=header)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="212427", end_color="212427", fill_type="solid")
            cell.border = thin_border

        for row_idx, asset in enumerate(assets_data, 2):
            size_mb = (asset.get("file_size_bytes", 0) or 0) / 1024 / 1024
            values = [
                asset.get("asset_type", "—"),
                asset.get("file_key", "—"),
                f"{size_mb:.1f}",
                asset.get("resolution_cm", "—"),
                asset.get("crs_epsg", "—"),
                "Sim" if asset.get("cog_validated") else "Não",
            ]
            for col, val in enumerate(values, 1):
                cell = ws_assets.cell(row=row_idx, column=col, value=str(val))
                cell.font = data_font
                cell.border = thin_border

        for col in range(1, len(asset_headers) + 1):
            ws_assets.column_dimensions[get_column_letter(col)].width = 20

        # Write to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_bytes = output.read()
        logger.info(f"Excel generated: {len(excel_bytes)} bytes")
        return excel_bytes

    def _generate_placeholder_pdf(self, project_data: dict) -> bytes:
        """Generate a minimal placeholder PDF when WeasyPrint is unavailable."""
        content = f"""
        <html>
        <body style="font-family: sans-serif; padding: 40px;">
            <h1>CM TECHMAP — Relatório</h1>
            <p>Projeto: {project_data.get('name', 'N/A')}</p>
            <p>Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
            <p><em>WeasyPrint não disponível — instale para relatórios completos.</em></p>
        </body>
        </html>
        """
        try:
            from weasyprint import HTML
            return HTML(string=content).write_pdf()
        except ImportError:
            # Return minimal valid PDF
            return b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\nxref\n0 1\ntrailer\n<<>>\nstartxref\n0\n%%EOF"
